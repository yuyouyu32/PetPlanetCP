from openai import OpenAI
from PetPlanetCP.src.cp_generation.promts.kuma import *
import random
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
import math
import time
import json
import requests


Client = OpenAI(api_key = "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")
N = 3
Cost = 0

def read_excel(filename: str):
    workbook = load_workbook(filename=filename, read_only=False, data_only=True)
    data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        merged_data = {}
        for merged_cell in sheet.merged_cells:
            min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_data[(row, col)] = top_left_cell_value

        data[sheet_name] = [[merged_data.get((row_idx+1, col_idx+1), cell.value)
                             for col_idx, cell in enumerate(row)] for row_idx, row in enumerate(sheet.iter_rows())]

    return data


# GPT-4
def calculate_token_price(prompt_tokens, completion_tokens):
    return 7.13 * (0.01 * prompt_tokens / 1000 + 0.03 * completion_tokens / 1000)

def get_copywriter_from_GPT(sys_prompt: str, user_prompt: str, n: int = 1):
    """
    Description: 
        Get copywriter from GPT-4
    Args:
        sys_prompt (str): System prompt
        user_prompt (str): User prompt
        n (int, optional): Number of copywriters to generate. Defaults to 1.
    Returns:
        copywriters (list): List of copywriters
        cost (float): Cost price(￥) of generating copywriters
    """
    response = Client.chat.completions.create(
    model="gpt-4-1106-preview",
    messages=[
        {"role": "system", "content": sys_prompt},
        {"role": "user", "content": user_prompt},
    ],
    n = n
    ).dict()
    cost = calculate_token_price(response['usage']['prompt_tokens'], response['usage']['completion_tokens'])
    copywriters = []
    for choice in range(n):
        copywriters.append(response['choices'][choice]['message']['content'])
    return copywriters, cost
# Baidu
def get_access_token():
    """
    使用 API Key，Secret Key 获取access_token，替换下列示例中的应用API Key、应用Secret Key
    """
        
    url = "https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=uEFcOQL51OpXMejaB5NLl2B5&client_secret=itckeuCsN5Ffq0Yy13EjhtImjvg1GVjt"
    
    payload = json.dumps("")
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }
    
    response = requests.request("POST", url, headers=headers, data=payload)
    return response.json().get("access_token")
    
def get_copywriter_from_Baidu(sys_prompt: str, user_prompt: str):
    url = "https://aip.baidubce.com/rpc/2.0/ai_custom/v1/wenxinworkshop/chat/completions_pro?access_token=" + get_access_token()
    
    payload = json.dumps({
        "messages": [
            {
                "role": "user",
                "content": user_prompt
            }
        ]
        ,
        "system": sys_prompt
    })
    headers = {
        'Content-Type': 'application/json'
    }
    
    response = requests.request("POST", url, headers=headers, data=payload)
    copyright = response.json().get("result")
    total_tokens = response.json().get("usage").get("total_tokens")
    cost = 0.12 * total_tokens / 1000
    return copyright, cost

if __name__ == '__main__':
    file_path = './各省特色整理_文案.xlsx'
    save_path = './各省特色整理_文案.xlsx'
    data = read_excel(file_path)
    dfs = {sheet_name: pd.DataFrame(sheet_data[1:], columns=sheet_data[0]) for sheet_name, sheet_data in data.items()}
    with pd.ExcelFile(file_path) as xls:
        sheet_names = xls.sheet_names


    for sheet_name in sheet_names:
        if sheet_name != '东线-风景篇': continue
        print(sheet_name, '开始获取文案...')
        sys_prompt = SystemPromptTemplate.format()
        content_prompt = ContentPromptTemplate.format(content_req=content_req)
        dataframe = dfs[sheet_name]
        for index, row in tqdm(dataframe.iterrows(), total=dataframe.shape[0]):
            province = row['省份']
            local = row['地点']
            name = row['名称']
            description = row['描述']
            if (isinstance(name, float) and math.isnan(name)) or (isinstance(name, str) and name.strip()) == '' or name is None:
                dataframe.loc[index, '诗句查询-百度'] = None
                continue
            if (isinstance(local, float) and math.isnan(local)) or (isinstance(local, str) and local.strip()) == '' or local is None:
                local = ''
            if (isinstance(province, float) and math.isnan(province)) or (isinstance(province, str) and province.strip()) == '' or province is None:
                province = ''
            title = province + local + name
            if (isinstance(description, float) and math.isnan(description)) or (isinstance(description, str) and description.strip()) == '' or description is None:
                description = ''
            copywriters = ""
            for i in range(2):
                random_examples = random.sample(examples, 5)
                examples_prompt = ExamplesPromptTemplate.format(examples="\n".join(random_examples))
                task_prompt = TaskPromptTemplate.format(title=title, description=description, content_prompt=content_prompt)
                results, cost = get_copywriter_from_Baidu(sys_prompt=sys_prompt, user_prompt=task_prompt)
                Cost += cost
                # copywriters += '\n\n'.join(results)
                copywriters += results
                if i == 0:
                    copywriters += '\n\n'
                time.sleep(random.randint(1, 3))
            dataframe.loc[index, '诗句查询-百度'] = copywriters
            # Save the dataframe to excel
            with pd.ExcelWriter(save_path) as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        print(sheet_name, 'Done!')
    print('Cost: ', Cost)
        